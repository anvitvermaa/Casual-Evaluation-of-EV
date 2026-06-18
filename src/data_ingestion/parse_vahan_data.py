"""
Data Ingestion Module: Vahan State-Level Panel Parser
Reads all state × year Excel files from data/raw/vehicle_registrations/states/,
extracts monthly EV and total registrations, and outputs a clean unified CSV
ready for DuckDB ingestion.

File naming convention expected: <state>_<year>.xlsx
e.g.: maharashtra_2022.xlsx, gujarat_2023.xlsx
"""

import os
import sys
import re
import glob
import openpyxl
import pandas as pd

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))
from config import settings

# All known Electric fuel row labels in Vahan
EV_FUEL_LABELS = {"ELECTRIC(BOV)", "STRONG HYBRID EV"}

# Month column order in the Vahan Excel layout
MONTH_COLS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
              "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]

# Map state filenames to clean standardized state names
STATE_NAME_MAP = {
    "maharashtra":    "MAHARASHTRA",
    "gujarat":        "GUJARAT",
    "karnataka":      "KARNATAKA",
    "tamilnadu":      "TAMIL NADU",
    "andhrapradesh":  "ANDHRA PRADESH",
    "telengana":      "TELANGANA",
    "madhyapradesh":  "MADHYA PRADESH",
    "rajasthan":      "RAJASTHAN",
    "uttarpradesh":   "UTTAR PRADESH",
}


def parse_excel_file(filepath: str, state_name: str, year: int) -> list[dict]:
    """
    Parse a single Vahan Excel file.
    Returns a list of dicts with keys: state, year_month, ev_registrations, total_registrations.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # --- Step 1: Find the header row containing month names (JAN, FEB, ...) ---
    header_row_idx = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        row_str = [str(c).strip().upper() if c else "" for c in row]
        if "JAN" in row_str and "FEB" in row_str:
            header_row_idx = row_idx
            # Build a col_index → month_name mapping
            month_col_map = {}
            for col_idx, cell_val in enumerate(row_str):
                if cell_val in MONTH_COLS:
                    month_col_map[col_idx] = cell_val
            break

    if header_row_idx is None:
        print(f"  [WARNING] Could not find month header row in {filepath}. Skipping.")
        return []

    # --- Step 2: Collect all data rows (rows below header) ---
    all_rows = list(ws.iter_rows(values_only=True))
    data_rows = all_rows[header_row_idx + 1:]

    # --- Step 3: Build monthly totals for ev and all fuels ---
    # ev_by_month[month_name] = sum of EV rows for that month
    ev_by_month    = {m: 0 for m in MONTH_COLS}
    total_by_month = {m: 0 for m in MONTH_COLS}

    for row in data_rows:
        # Fuel label is typically in column index 1
        if len(row) < 3:
            continue
        fuel_label = str(row[1]).strip().upper() if row[1] else ""

        # Strip non-breaking spaces and whitespace
        fuel_label = re.sub(r"\s+", " ", fuel_label.replace("\xa0", " ")).strip()

        if not fuel_label or fuel_label in ("", "FUEL", "S NO"):
            continue

        for col_idx, month_name in month_col_map.items():
            if col_idx >= len(row):
                continue
            raw_val = row[col_idx]
            if raw_val is None:
                continue
            # Vahan stores numbers as strings with commas e.g. "1,35,985"
            try:
                val = int(str(raw_val).replace(",", "").strip())
            except ValueError:
                continue

            total_by_month[month_name] += val
            if fuel_label in EV_FUEL_LABELS:
                ev_by_month[month_name] += val

    # --- Step 4: Convert to row-per-month records ---
    records = []
    for month_name in MONTH_COLS:
        month_num = MONTH_COLS.index(month_name) + 1
        year_month = f"{year}-{month_num:02d}"

        # For 2026 files the dashboard only has Jan–Jun; skip zero months
        if total_by_month[month_name] == 0:
            continue

        records.append({
            "state":               state_name,
            "year_month":          year_month,
            "ev_registrations":    ev_by_month[month_name],
            "total_registrations": total_by_month[month_name],
        })

    return records


def parse_all_state_files():
    """Main entry point: parse all 45 state files and write a unified CSV."""
    states_dir = os.path.join(settings.RAW_DATA_DIR, "vehicle_registrations", "states")
    output_path = os.path.join(settings.RAW_DATA_DIR, "vehicle_registrations", "vahan_state_panel.csv")

    all_files = sorted(glob.glob(os.path.join(states_dir, "*.xlsx")))
    if not all_files:
        print(f"[ERROR] No .xlsx files found in {states_dir}")
        return

    print(f"Found {len(all_files)} files. Parsing...")

    all_records = []
    for filepath in all_files:
        filename = os.path.basename(filepath).replace(".xlsx", "")
        parts = filename.rsplit("_", 1)       # split on the LAST underscore only
        if len(parts) != 2:
            print(f"  [SKIP] Unrecognised filename format: {filename}")
            continue

        state_key, year_str = parts[0], parts[1]
        if state_key not in STATE_NAME_MAP:
            print(f"  [SKIP] Unknown state key '{state_key}' in {filename}")
            continue
        try:
            year = int(year_str)
        except ValueError:
            print(f"  [SKIP] Cannot parse year from '{year_str}' in {filename}")
            continue

        state_name = STATE_NAME_MAP[state_key]
        print(f"  Parsing {filename}  →  state={state_name}, year={year}")

        try:
            records = parse_excel_file(filepath, state_name, year)
            all_records.extend(records)
            print(f"    → {len(records)} monthly rows extracted.")
        except Exception as e:
            print(f"    [ERROR] Failed to parse {filename}: {e}")

    if not all_records:
        print("[ERROR] No records extracted. Check file formats and paths.")
        return

    df = pd.DataFrame(all_records)
    df["ev_penetration_rate"] = df["ev_registrations"] / df["total_registrations"] * 100
    df = df.sort_values(["state", "year_month"]).reset_index(drop=True)

    df.to_csv(output_path, index=False)
    print(f"\n✅ Panel CSV written → {output_path}")
    print(f"   Total rows : {len(df)}")
    print(f"   States     : {sorted(df['state'].unique())}")
    print(f"   Date range : {df['year_month'].min()} → {df['year_month'].max()}")
    print(f"\nSample rows:")
    print(df.head(15).to_string(index=False))


if __name__ == "__main__":
    parse_all_state_files()
